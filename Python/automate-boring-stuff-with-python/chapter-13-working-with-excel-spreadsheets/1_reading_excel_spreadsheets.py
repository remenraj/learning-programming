# The OpenPyXL third-party module handles Excel spreadsheets (.xlsx files).
# openpyxl.load_workbook(filename) returns a Workbook object.
# get_sheet_names() and get_sheet_by_name() help get Worksheet objects.
# The square brackets in sheet[‘A1'] get Cell objects.
# Cell objects have a "value" member variable with the content of that cell.
# The cell() method also returns a Cell object from a sheet.

# install openpyxl: pip install openpyxl

import openpyxl, os

# change current working directory to where the file is located
# print(os.path.dirname(__file__))    # prints the path to the file
# print(os.getcwd())    # get current working directory
os.chdir(os.path.dirname(__file__))

workbook = openpyxl.load_workbook("example.xlsx")
print(type(workbook))

# get sheet names
sheet_names = workbook.sheetnames
print(sheet_names)

# sheet1 object
sheet_1 = workbook["Sheet1"]
print(type(sheet_1))

print(sheet_1["A1"])  # cell object
print(sheet_1["A1"].value)  # cell value
print(type(sheet_1["A1"].value))  # type of cell value
print(str(sheet_1["A1"].value))  # string representation of cell value

print(sheet_1["B1"].value)  # cell value

print(sheet_1.cell(row=1, column=3).value)  # cell value