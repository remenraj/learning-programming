# You can view and modify a sheet's name with its "title" member variable.
# Changing a cell's value is done using the square brackets, just like changing a value in a list or dictionary.
# Changes you make to the workbook object can be saved with the save() method.

import openpyxl, os


# create a new workbook
wb = openpyxl.Workbook()

print(wb.sheetnames)

sheet = wb["Sheet"]

print(sheet["A1"].value == None)

sheet["A1"] = "Hello"

# change directory
os.chdir(os.path.dirname(__file__))

# save the workbook
wb.save("example2.xlsx")

# create new sheet
sheet2 = wb.create_sheet(title="My new sheet name") # also takes index=0/1/2/3.. to create a new sheet at the beginning
print(wb.sheetnames)

wb.save("example3.xlsx")